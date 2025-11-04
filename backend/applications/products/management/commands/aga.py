# applications/products/management/commands/aga.py
from __future__ import annotations

from typing import Dict, List, Tuple, Optional
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import F, Value, Case, When, IntegerField

from applications.products.models import Product

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class Command(BaseCommand):
    help = (
        "Синхронизировать остатки.\n"
        "  Без --file: stock_qty = available_stock_qty (массовый UPDATE).\n"
        "  С --file: взять 1-ю колонку как имя товара, 2-ю колонку как количество и обновить stock_qty."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", help="Путь к .xlsx файлу.")
        parser.add_argument("--sheet", help="Имя листа (по умолчанию активный).")
        parser.add_argument(
            "--header-row",
            type=int,
            help="Номер строки (1-based) заголовков. Если указан — эта строка будет пропущена."
        )
        parser.add_argument("--dry-run", action="store_true", help="Только показать, что обновится.")

        # Режим без Excel
        parser.add_argument(
            "--only-non-equal",
            action="store_true",
            help="(Без Excel) Обновлять только где stock_qty != available_stock_qty.",
        )

    def handle(self, *args, **options):
        file_path = options.get("file")
        if not file_path:
            return self._handle_simple_sync(options)

        if load_workbook is None:
            raise CommandError("Нужен openpyxl: pip install openpyxl")

        return self._handle_excel_mode(options)

    # ---------- режим без Excel ----------
    def _handle_simple_sync(self, options) -> None:
        qs = Product.objects.all()
        if options["only_non_equal"]:
            qs = qs.exclude(stock_qty=F("available_stock_qty"))

        count = qs.count()
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING(f"[DRY RUN] Будет обновлено: {count} товар(ов)."))
            return

        with transaction.atomic():
            updated = qs.update(stock_qty=F("available_stock_qty"))
        self.stdout.write(self.style.SUCCESS(f"Готово. Обновлено {updated} товар(ов)."))

    # ---------- режим с Excel ----------
    def _handle_excel_mode(self, options) -> None:
        file_path: str = options["file"]
        sheet_name: Optional[str] = options.get("sheet")
        header_row_opt: Optional[int] = options.get("header-row")
        dry_run: bool = options["dry_run"]

        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # стартовая строка для данных: если задан header_row — начинаем со следующей
        start_row = (header_row_opt + 1) if header_row_opt else 1

        excel_map, bad_rows = self._read_rows_positional(ws, start_row)

        if not excel_map:
            raise CommandError(
                "В Excel нет валидных строк (1-я колонка — имя, 2-я — количество)."
            )

        names = list(excel_map.keys())
        existing = Product.objects.filter(name__in=names).values_list("name", flat=True)
        existing_set = set(existing)
        missing_in_db = sorted(set(names) - existing_set)

        whens = [When(name=n, then=Value(excel_map[n])) for n in existing_set]
        case_expr = Case(*whens, output_field=IntegerField())

        if dry_run:
            self.stdout.write(self.style.WARNING("[DRY RUN]"))
            self.stdout.write(f"  Позиции в Excel: {len(excel_map)}")
            self.stdout.write(f"  Найдено в БД:   {len(existing_set)}")
            if missing_in_db:
                self.stdout.write(self.style.WARNING("  Нет в БД (по имени):"))
                for n in missing_in_db[:50]:
                    self.stdout.write(f"    - {n}")
                if len(missing_in_db) > 50:
                    self.stdout.write(f"    ... и ещё {len(missing_in_db) - 50}")
            if bad_rows:
                self.stdout.write(self.style.WARNING("  Строки с ошибками:"))
                for r_idx, err in bad_rows[:20]:
                    self.stdout.write(f"    {r_idx}: {err}")
                if len(bad_rows) > 20:
                    self.stdout.write(f"    ... и ещё {len(bad_rows) - 20}")
            return

        updated = 0
        with transaction.atomic():
            if existing_set:
                updated = (
                    Product.objects
                    .filter(name__in=existing_set)
                    .update(stock_qty=case_expr)  # <— обновляем только stock_qty
                )

        self.stdout.write(self.style.SUCCESS(f"Готово. Обновлено {updated} товар(ов)."))
        if missing_in_db:
            self.stdout.write(self.style.WARNING("Не найдены в БД (по точному имени):"))
            for n in missing_in_db[:100]:
                self.stdout.write(f"  - {n}")
            if len(missing_in_db) > 100:
                self.stdout.write(f"  ... и ещё {len(missing_in_db) - 100}")
        if bad_rows:
            self.stdout.write(self.style.WARNING("Строки с ошибками:"))
            for r_idx, err in bad_rows[:20]:
                self.stdout.write(f"  {r_idx}: {err}")
            if len(bad_rows) > 20:
                self.stdout.write(f"  ... и ещё {len(bad_rows) - 20}")

    # === helpers ===

    def _read_rows_positional(
        self, ws, start_row: int
    ) -> tuple[Dict[str, int], List[Tuple[int, str]]]:
        """
        Читает строки, начиная с start_row.
        Ожидается, что:
          - 1-я колонка: имя товара
          - 2-я колонка: количество (записываем в stock_qty)
        Пустые имена пропускаются. Кол-во приводим к int, отрицательные -> 0.
        """
        excel_map: Dict[str, int] = {}
        bad_rows: List[Tuple[int, str]] = []

        for r_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # гарантируем наличие хотя бы двух колонок
            name_raw = row[0] if len(row) >= 1 else None
            qty_raw = row[1] if len(row) >= 2 else None

            # пропуск пустых имён
            if name_raw is None:
                continue
            name = str(name_raw).strip()
            if not name:
                continue

            # пропуск пустых количеств
            if qty_raw is None or str(qty_raw).strip() == "":
                continue

            # парсинг количества
            try:
                qty = int(float(str(qty_raw).replace(",", ".").strip()))
                if qty < 0:
                    qty = 0
            except Exception:
                bad_rows.append((r_idx, f"Некорректное число во 2-й колонке: {qty_raw!r}"))
                continue

            excel_map[name] = qty

        return excel_map, bad_rows

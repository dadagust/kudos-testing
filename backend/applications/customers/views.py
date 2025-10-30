"""REST API endpoints for customer domain."""

from __future__ import annotations

from io import BytesIO

from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Contact, Customer, PhoneNormalizer
from .permissions import CustomerAccessPolicy
from .serializers import (
    ContactSerializer,
    CustomerDetailSerializer,
    CustomerWriteSerializer,
    get_customer_serializer,
)
from .utils import QueryParamsHelper


class CustomerViewSet(viewsets.ModelViewSet):
    """Full CRUD endpoint for customers with scoped access."""

    queryset = (
        Customer.objects.all()
        .select_related('company', 'owner')
        .prefetch_related(
            Prefetch('contacts', queryset=Contact.objects.filter(is_active=True)),
        )
    )
    permission_classes = [IsAuthenticated, CustomerAccessPolicy]

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(is_active=True)

        user = self.request.user

        if not user or not user.is_authenticated:
            return queryset.none()

        if user.is_staff:
            return queryset

        if user.has_perm('customers.view_customer'):
            return queryset.filter(owner=user)

        return queryset.none()

    def get_serializer_class(self):
        if self.action in {'create', 'update', 'partial_update'}:
            return CustomerWriteSerializer
        return get_customer_serializer(self.action)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def filter_queryset(self, queryset):  # type: ignore[override]
        helper = QueryParamsHelper(self.request)
        queryset = queryset

        search = helper.get_search()
        if search:
            queryset = queryset.filter(
                Q(display_name__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
                | Q(phone__icontains=search)
            )

        email = helper.get_filter('email').lower()
        if email:
            queryset = queryset.filter(email__iexact=email)

        phone = helper.get_filter('phone')
        if phone:
            phone_normalized = PhoneNormalizer.normalize(phone)
            queryset = queryset.filter(phone_normalized=phone_normalized)

        company_id = helper.get_filter('company_id')
        if company_id:
            queryset = queryset.filter(company_id=company_id)

        created_from = helper.get_filter_datetime('created', 'from')
        if created_from:
            queryset = queryset.filter(created__gte=created_from)
        created_to = helper.get_filter_datetime('created', 'to')
        if created_to:
            queryset = queryset.filter(created__lte=created_to)

        sort_fields = helper.get_sort_fields({'name': 'display_name', 'created': 'created'})
        if sort_fields:
            queryset = queryset.order_by(*sort_fields)
        else:
            queryset = queryset.order_by('-created')

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = CustomerWriteSerializer(
            data=request.data, context=self.get_serializer_context()
        )
        serializer.is_valid(raise_exception=True)
        customer = serializer.save()
        read_serializer = CustomerDetailSerializer(customer, context=self.get_serializer_context())
        payload = {'data': read_serializer.data}
        headers = self.get_success_headers(read_serializer.data)
        return Response(payload, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = CustomerWriteSerializer(
            instance, data=request.data, context=self.get_serializer_context(), partial=partial
        )
        serializer.is_valid(raise_exception=True)
        customer = serializer.save()
        read_serializer = CustomerDetailSerializer(customer, context=self.get_serializer_context())
        return Response({'data': read_serializer.data})

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = CustomerDetailSerializer(instance, context=self.get_serializer_context())
        return Response({'data': serializer.data})

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if isinstance(response.data, list):
            response.data = {'data': response.data}
        return response

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Клиенты'

        headers = [
            'ID',
            'Имя',
            'Тип',
            'Email',
            'Телефон',
            'Компания',
            'Ответственный',
            'Согласие на обработку',
            'Дата создания',
            'Дата обновления',
            'Заметки',
        ]
        worksheet.append(headers)

        def format_datetime(value):
            if not value:
                return ''
            localized = timezone.localtime(value)
            return localized.strftime('%d.%m.%Y %H:%M')

        for customer in queryset:
            worksheet.append(
                [
                    str(customer.id),
                    customer.full_name,
                    customer.get_customer_type_display(),
                    customer.email,
                    customer.phone,
                    customer.company.name if customer.company else '',
                    customer.owner.email if customer.owner else '',
                    'Да' if customer.gdpr_consent else 'Нет',
                    format_datetime(customer.created),
                    format_datetime(customer.modified),
                    customer.notes,
                ]
            )

        for column_cells in worksheet.iter_cols(min_row=1, max_row=worksheet.max_row):
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            adjusted_width = min(max(max_length + 2, 12), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime('%Y%m%d-%H%M%S')
        filename = f'customers-{timestamp}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def perform_destroy(self, instance: Customer) -> None:
        instance.is_active = False
        instance.save(update_fields=['is_active', 'modified'])


    @action(detail=True, methods=['get', 'post'], url_path='contact')
    def contact(self, request, *args, **kwargs):
        customer = self.get_object()
        queryset = customer.contacts.filter(is_active=True).order_by('-is_primary', '-created')

        if request.method == 'GET':
            serializer = ContactSerializer(
                queryset,
                many=True,
                context={
                    'customer': customer,
                    'request': request,
                },
            )
            return Response({'data': serializer.data})

        serializer = ContactSerializer(
            data=request.data,
            context={
                'customer': customer,
                'request': request,
            },
        )
        serializer.is_valid(raise_exception=True)
        contact = serializer.save()
        read_serializer = ContactSerializer(
            contact,
            context={
                'customer': customer,
                'request': request,
            },
        )
        headers = self.get_success_headers(read_serializer.data)
        return Response(
            {'data': read_serializer.data},
            status=status.HTTP_201_CREATED,
            headers=headers,
        )


__all__ = ['CustomerViewSet']

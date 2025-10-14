from __future__ import annotations

from io import BytesIO

from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from applications.users.models import RoleChoices

from .models import Contact, Customer


class CustomerAPITests(APITestCase):
    def setUp(self):
        self.User = get_user_model()
        self.manager = self.User.objects.create_user(
            username='manager@example.com',
            email='manager@example.com',
            password='ChangeMe123!',
        )
        self.manager.profile.role = RoleChoices.SALES_MANAGER
        self.manager.profile.save()

        self.customer_user = self.User.objects.create_user(
            username='client@example.com',
            email='client@example.com',
            password='ChangeMe123!',
        )
        self.customer_user.profile.role = RoleChoices.CUSTOMER
        self.customer_user.profile.save()

        self.list_url = '/api/v1/customer/'

    def authenticate(self, user):
        self.client.force_authenticate(user=user)

    def test_sales_manager_can_create_customer(self):
        self.authenticate(self.manager)
        payload = {
            'customer_type': 'personal',
            'first_name': 'Иван',
            'last_name': 'Петров',
            'email': 'Test@Example.com',
            'phone': '+7 (999) 123-45-67',
            'gdpr_consent': True,
            'notes': 'Новый клиент',
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data = response.json()['data']
        self.assertEqual(data['email'], 'test@example.com')
        self.assertEqual(data['phone'], '+79991234567')
        self.assertTrue(Customer.objects.filter(email='test@example.com').exists())

    def test_empty_customer_list_returns_success(self):
        self.authenticate(self.manager)
        response = self.client.get(self.list_url, {'page': 1, 'page_size': 10})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload['data'], [])
        self.assertEqual(payload['meta']['pagination']['total_items'], 0)
        self.assertEqual(payload['meta']['pagination']['page'], 1)

    def test_customer_sees_only_own_record(self):
        self.authenticate(self.manager)
        self.client.post(
            self.list_url,
            {
                'customer_type': 'personal',
                'first_name': 'Alice',
                'email': 'alice@example.com',
                'phone': '+7 900 111-22-33',
            },
            format='json',
        )
        self.client.post(
            self.list_url,
            {
                'customer_type': 'personal',
                'first_name': 'Bob',
                'email': 'bob@example.com',
                'phone': '+7 900 444-55-66',
                'owner_id': self.customer_user.id,
            },
            format='json',
        ).json()['data']

        self.client.force_authenticate(user=None)
        self.authenticate(self.customer_user)
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(len(payload['data']), 1)
        self.assertEqual(payload['data'][0]['email'], 'bob@example.com')

    def test_content_manager_cannot_access(self):
        content_user = self.User.objects.create_user(
            username='content@example.com',
            email='content@example.com',
            password='ChangeMe123!',
        )
        content_user.profile.role = RoleChoices.CONTENT_MANAGER
        content_user.profile.save()

        self.authenticate(content_user)
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_soft_delete_marks_inactive(self):
        self.authenticate(self.manager)
        created = self.client.post(
            self.list_url,
            {'customer_type': 'personal', 'first_name': 'Mark', 'phone': '+7 905 000-00-00'},
            format='json',
        ).json()['data']
        customer = Customer.objects.get(pk=created['id'])
        delete_response = self.client.delete(f"{self.list_url}{customer.id}/")
        self.assertEqual(delete_response.status_code, status.HTTP_204_NO_CONTENT)
        customer.refresh_from_db()
        self.assertFalse(customer.is_active)

    def test_can_create_contact_for_customer(self):
        self.authenticate(self.manager)
        created = self.client.post(
            self.list_url,
            {'customer_type': 'business', 'first_name': 'Corp', 'email': 'corp@example.com'},
            format='json',
        ).json()['data']
        customer_id = created['id']
        contacts_url = f"{self.list_url}{customer_id}/contact/"
        response = self.client.post(
            contacts_url,
            {'first_name': 'Анна', 'last_name': 'Иванова', 'email': 'anna@example.com'},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Contact.objects.filter(customer_id=customer_id).count(), 1)

        list_response = self.client.get(contacts_url)
        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(list_response.json()['data']), 1)

    def test_export_customers_respects_search_and_sort(self):
        self.authenticate(self.manager)
        payloads = [
            {
                'customer_type': 'personal',
                'first_name': 'Alice',
                'last_name': 'Zeta',
                'display_name': 'Alice Client',
                'email': 'alice@example.com',
            },
            {
                'customer_type': 'personal',
                'first_name': 'Bob',
                'last_name': 'Yellow',
                'display_name': 'Bob Client',
                'email': 'bob@example.com',
            },
            {
                'customer_type': 'business',
                'first_name': 'Charlie',
                'last_name': 'Xavier',
                'display_name': 'Charlie Client',
                'email': 'charlie@example.com',
            },
        ]

        for payload in payloads:
            response = self.client.post(self.list_url, payload, format='json')
            self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        export_url = f'{self.list_url}export/'
        response = self.client.get(export_url, {'search': 'client', 'sort': 'name'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 1)

        header = rows[0]
        name_index = header.index('Имя')

        exported_names = [row[name_index] for row in rows[1:]]
        self.assertEqual(exported_names, ['Alice Client', 'Bob Client', 'Charlie Client'])

        filtered_response = self.client.get(export_url, {'search': 'Bob'})
        self.assertEqual(filtered_response.status_code, status.HTTP_200_OK)
        workbook_filtered = load_workbook(BytesIO(filtered_response.content))
        sheet_filtered = workbook_filtered.active
        filtered_rows = list(sheet_filtered.iter_rows(values_only=True))
        self.assertEqual(len(filtered_rows) - 1, 1)
        self.assertEqual(filtered_rows[1][name_index], 'Bob Client')
